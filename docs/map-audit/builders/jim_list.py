"""Jim's own Keep/Delete decisions, which are the authority for this proposal.

Everything before this was inference: a billable-labour floor, then an invoice
test. Both were better than what came before and both were still wrong. The
floor could not tell a lost pursuit from a won job, because a pursuit carries
proposal time. The invoice test got closer but still kept `4th & Ash` (two
invoices) and `Courthouse North Block` (three, "lost to GS"), which Jim marks
Delete. There is no derivable signal for "did we get the job" -- Deltek has no
won/lost flag, and invoicing proposal work blurs the one proxy that looked
clean.

Jim went through all 75 rows on 10 Aug and marked each Keep or Delete with a
reason. That file is the source of truth here. It is read directly rather than
transcribed, so the proposal cannot drift from what he actually decided.

His 12 Aug call supersedes the email where they differ: downtown core only, and
the QR points at the live map rather than his bio.

The PBD column is his own shortlist for THIS pursuit (Park & Broadway).
"""
import io, json, os, re

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = r"C:/Users/ilalonde/Desktop/Jim/KOR-CA-Portfolio-Projects JD.xlsx"

# The downtown core, as used for the proposal figure.
DOWNTOWN = (32.703, 32.728, -117.176, -117.145)

KOR_CAT = "kor structural"
PRIOR_CAT = "jim desroches"


def rows():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["CA Portfolio"]
    head = None
    out = []
    adding = False
    for r in ws.iter_rows(values_only=True):
        vals = ["" if c is None else str(c).strip() for c in r]
        if not any(vals):
            continue
        if vals[0].lower().startswith("project") and "category" in vals[1].lower():
            head = vals
            continue
        if vals[0].lower().startswith("projects  to add") or vals[0].lower().startswith("projects to add"):
            adding = True
            continue
        if head is None or vals[0].lower().startswith(("highlighted rows", "some addresses", "kor structural \u2013", "75 projects")):
            continue
        d = dict(zip(head, vals))
        d["_add"] = adding
        out.append(d)
    return out


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    data = rows()
    keep, delete = [], []
    for d in data:
        decision = (d.get("Keep/Delete") or "").strip().lower()
        rec = {
            "name": d.get("Project", "").strip(),
            "category": (d.get("Category") or "").strip(),
            "developer": (d.get("Developer / firm") or "").strip(),
            "architect": (d.get("Architect") or "").strip(),
            "address": (d.get("Address") or "").strip(),
            "region": (d.get("Region") or "").strip(),
            "lat": num(d.get("Latitude")),
            "lng": num(d.get("Longitude")),
            "comment": (d.get("Comments by JD") or "").strip(),
            "pbd": bool((d.get("PBD") or "").strip()),
            "added": d["_add"],
        }
        (keep if decision == "keep" else delete).append(rec)

    lat0, lat1, lng0, lng1 = DOWNTOWN

    def downtown(r):
        return (r["lat"] is not None and r["lng"] is not None
                and lat0 <= r["lat"] <= lat1 and lng0 <= r["lng"] <= lng1)

    def is_kor(r):
        return r["category"].lower().startswith(KOR_CAT)

    def is_prior(r):
        return PRIOR_CAT in r["category"].lower()

    kor_dt = sorted((r for r in keep if is_kor(r) and downtown(r)), key=lambda r: r["name"].lower())
    prior_dt = sorted((r for r in keep if is_prior(r) and downtown(r)), key=lambda r: r["name"].lower())
    kor_sd = [r for r in keep if is_kor(r) and r["region"].lower() == "san diego"]
    no_coords = [r for r in keep if r["lat"] is None and r["region"].lower() == "san diego"]

    print("Jim marked %d Keep, %d Delete\n" % (len(keep), len(delete)))
    print("=== KOR, downtown core (%d) ===" % len(kor_dt))
    for r in kor_dt:
        print("   %-34s %-22s %s%s" % (r["name"][:34], r["developer"][:22], r["comment"][:40],
                                       "   [PBD]" if r["pbd"] else ""))
    print("\n=== Jim, prior to KOR, downtown core (%d) ===" % len(prior_dt))
    for r in prior_dt:
        print("   %-34s %s%s" % (r["name"][:34], r["comment"][:44], "   [PBD]" if r["pbd"] else ""))
    print("\n=== KOR across Greater San Diego (%d) ===" % len(kor_sd))
    print("\n=== DELETED by Jim (%d) ===" % len(delete))
    for r in sorted(delete, key=lambda r: r["name"].lower()):
        print("   %-34s %s" % (r["name"][:34], r["comment"][:52]))
    if no_coords:
        print("\n=== Keep, but no coordinates yet -- needs geocoding (%d) ===" % len(no_coords))
        for r in no_coords:
            print("   %-34s %s" % (r["name"][:34], r["address"][:44]))

    io.open(HERE + "/jim_list.json", "w", encoding="utf-8").write(json.dumps(
        {"kor_downtown": kor_dt, "prior_downtown": prior_dt,
         "kor_san_diego": kor_sd, "deleted": delete, "needs_geocode": no_coords},
        indent=1))
    print("\nwrote jim_list.json")


if __name__ == "__main__":
    main()
