#!/usr/bin/env python3
"""
Build a review workbook for recovered CSA A23.3 equation/figure blocks.

The source PDF is obfuscated, so the extracted image/PDF block is treated as
the source of truth. This workbook embeds the recovered block image beside
blank transcription fields for human-reviewed plain text, LaTeX, and optional
Excel-ready expressions.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PilImage


ROOT = Path(__file__).resolve().parents[1]
EXTRACT_DIR = ROOT / "docs" / "csa-a233-equation-extract"
MANIFEST = EXTRACT_DIR / "manifest.csv"
OUT_XLSX = EXTRACT_DIR / "csa-a233-equation-review.xlsx"


HEADERS = [
    "EquationId",
    "Page",
    "Block",
    "OriginalBlockImage",
    "ImagePath",
    "PdfPath",
    "BlockSizePt",
    "Status",
    "SplitNeeded",
    "PlainText",
    "LaTeX",
    "ExcelExpression",
    "Variables/Notes",
    "Reviewer",
]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def load_manifest() -> list[dict[str, str]]:
    if not MANIFEST.exists():
        raise FileNotFoundError(f"Missing manifest: {MANIFEST}")
    with MANIFEST.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def add_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("How to use")
    ws["A1"] = "CSA A23.3 equation extraction review workflow"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Purpose"
    ws["B3"] = "Preserve each original extracted block and transcribe equations beside it. The image/PDF block is the source of truth."
    ws["A5"] = "Status values"
    ws["B5"] = "Needs review, Transcribed, Verified, Not equation, Needs split"
    ws["A7"] = "Recommended transcription"
    ws["B7"] = "Use LaTeX for correctness. PlainText is for search. ExcelExpression is only for equations you want calculable."
    ws["A9"] = "When a block contains multiple equations"
    ws["B9"] = "Set SplitNeeded = Yes, duplicate the row manually, and transcribe one equation per row while keeping the same image reference."
    ws["A11"] = "Correctness rule"
    ws["B11"] = "Do not mark Verified until a human has checked every symbol, subscript, superscript, inequality, and coefficient against the image."
    ws["A13"] = "Why no automatic OCR text is prefilled"
    ws["B13"] = "The PDF text layer is obfuscated and math OCR is not installed locally. Prefilling guessed math would create false confidence."
    for row in ws.iter_rows(min_row=1, max_row=13, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110


def build() -> None:
    rows = load_manifest()
    wb = Workbook()
    ws = wb.active
    ws.title = "Equation review"
    add_guide_sheet(wb)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_validation = DataValidation(
        type="list",
        formula1='"Needs review,Transcribed,Verified,Not equation,Needs split"',
        allow_blank=False,
    )
    split_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(split_validation)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}1"

    for r_idx, row in enumerate(rows, start=2):
        page = int(row["page"])
        block = row["xobject"].strip("/")
        eq_id = f"CSA-A23.3-24-p{page:03d}-{block}"
        png_path = ROOT / row["png"]
        pdf_path = ROOT / row["pdf"]
        size_pt = f'{row["bbox_width"]} x {row["bbox_height"]}'

        values = [
            eq_id,
            page,
            block,
            None,
            rel(png_path),
            rel(pdf_path),
            size_pt,
            "Needs review",
            "No",
            "",
            "",
            "",
            "",
            "",
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.cell(row=r_idx, column=5).hyperlink = rel(png_path)
        ws.cell(row=r_idx, column=5).style = "Hyperlink"
        ws.cell(row=r_idx, column=6).hyperlink = rel(pdf_path)
        ws.cell(row=r_idx, column=6).style = "Hyperlink"
        status_validation.add(ws.cell(row=r_idx, column=8))
        split_validation.add(ws.cell(row=r_idx, column=9))

        if png_path.exists():
            with PilImage.open(png_path) as img:
                width, height = img.size
            max_width = 420
            max_height = 130
            scale = min(max_width / width, max_height / height, 1.0)
            ximg = XlsxImage(str(png_path))
            ximg.width = int(width * scale)
            ximg.height = int(height * scale)
            ws.add_image(ximg, f"D{r_idx}")
            ws.row_dimensions[r_idx].height = max(78, ximg.height * 0.78)

    widths = {
        "A": 28,
        "B": 8,
        "C": 10,
        "D": 58,
        "E": 54,
        "F": 54,
        "G": 16,
        "H": 16,
        "I": 12,
        "J": 42,
        "K": 56,
        "L": 42,
        "M": 48,
        "N": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws["J1"].comment = Comment("Search-friendly transcription. Keep it simple; use LaTeX for exact math.", "Codex")
    ws["K1"].comment = Comment("Preferred exact transcription format for engineering equations.", "Codex")
    ws["L1"].comment = Comment("Optional calculable Excel form, only where useful.", "Codex")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    build()
