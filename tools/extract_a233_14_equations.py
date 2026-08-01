#!/usr/bin/env python3
"""
Extract equation-image candidates from A23.3-14 Design of Concrete Structures.

This PDF has a usable text layer, but displayed equations are embedded as
separate small images. This script uses Poppler's pdfimages to extract the
embedded images intact, then builds a review workbook with each image, page,
nearby equation labels from the text layer, and blank human-review transcription
columns.
"""

from __future__ import annotations

import csv
import re
import shutil
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PilImage


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PDF = ROOT / "docs" / "A23.3-14 Design of Concrete Structures.pdf"
OUT_DIR = ROOT / "docs" / "a233-14-equation-extract"
IMG_DIR = OUT_DIR / "images"
MANIFEST = OUT_DIR / "manifest.csv"
WORKBOOK = OUT_DIR / "a233-14-equation-review.xlsx"


def run(args: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        args,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def parse_pdfimages_list() -> list[dict[str, object]]:
    result = run(["pdfimages", "-list", str(SOURCE_PDF)])
    rows: list[dict[str, object]] = []
    for line in result.stdout.splitlines():
        if not re.match(r"\s*\d+\s+\d+\s+image\s+", line):
            continue
        parts = line.split()
        rows.append(
            {
                "page": int(parts[0]),
                "num": int(parts[1]),
                "width": int(parts[3]),
                "height": int(parts[4]),
                "color": parts[5],
                "components": parts[6],
                "bits_per_component": parts[7],
                "encoding": parts[8],
                "object_id": parts[10] if len(parts) > 10 else "",
            }
        )
    return rows


def extract_raw_images() -> dict[int, Path]:
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    prefix = IMG_DIR / "raw"
    subprocess.run(
        ["pdfimages", "-png", str(SOURCE_PDF), str(prefix)],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    by_num: dict[int, Path] = {}
    for path in IMG_DIR.glob("raw-*.*"):
        match = re.match(r"raw-(\d+)\.", path.name)
        if match:
            by_num[int(match.group(1))] = path
    return by_num


def extract_page_text(page: int) -> str:
    result = run(["pdftotext", "-f", str(page), "-l", str(page), "-layout", str(SOURCE_PDF), "-"])
    text = result.stdout
    # Strip licence/footer noise and compact whitespace for context snippets.
    cleaned_lines = []
    for line in text.splitlines():
        if "Licensed for/Autorisé" in line:
            continue
        if "Single user license" in line:
            continue
        if "© 2014 CSA Group" in line:
            continue
        cleaned_lines.append(line.rstrip())
    return "\n".join(cleaned_lines).strip()


def page_equation_labels(text: str) -> list[str]:
    labels: list[str] = []
    seen: set[str] = set()
    for line in text.splitlines():
        stripped = line.strip()
        match = re.fullmatch(r"Equation\s+([A-Z]?\d+(?:\.\d+)*[a-z]?)", stripped)
        if not match:
            continue
        label = f"Equation {match.group(1)}"
        if label not in seen:
            labels.append(label)
            seen.add(label)
    return labels


def context_around_labels(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    snippets: list[str] = []
    for i, line in enumerate(lines):
        if "Equation " not in line:
            continue
        start = max(0, i - 4)
        end = min(len(lines), i + 2)
        snippet = " ".join(l for l in lines[start:end] if l)
        snippets.append(re.sub(r"\s+", " ", snippet))
    return " || ".join(snippets)[:1200]


def candidate_kind(width: int, height: int) -> str:
    if height > 220 or width > 1200:
        return "Likely figure/table"
    if height <= 190:
        return "Equation candidate"
    return "Review"


def build_outputs() -> None:
    if not SOURCE_PDF.exists():
        raise FileNotFoundError(SOURCE_PDF)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    IMG_DIR.mkdir(parents=True, exist_ok=True)

    image_rows = parse_pdfimages_list()
    raw_images = extract_raw_images()

    text_cache: dict[int, str] = {}
    per_page_counter: defaultdict[int, int] = defaultdict(int)
    per_page_total: defaultdict[int, int] = defaultdict(int)
    manifest_rows: list[dict[str, object]] = []

    for row in image_rows:
        page = int(row["page"])
        num = int(row["num"])
        if page == 1 and num == 0:
            continue
        if num in raw_images:
            per_page_total[page] += 1

    for row in image_rows:
        page = int(row["page"])
        num = int(row["num"])
        # Skip the cover-page full-page image. It is not an equation.
        if page == 1 and num == 0:
            continue
        src = raw_images.get(num)
        if src is None:
            continue

        per_page_counter[page] += 1
        stem = f"eqimg_p{page:03d}_{per_page_counter[page]:02d}_img{num:03d}"
        dest = IMG_DIR / f"{stem}{src.suffix.lower()}"
        if src.resolve() != dest.resolve():
            shutil.copyfile(src, dest)

        if page not in text_cache:
            text_cache[page] = extract_page_text(page)
        page_text = text_cache[page]
        labels = page_equation_labels(page_text)
        assigned_label = labels[per_page_counter[page] - 1] if len(labels) == per_page_total[page] and per_page_counter[page] <= len(labels) else ""
        kind = candidate_kind(int(row["width"]), int(row["height"]))

        manifest_rows.append(
            {
                "EquationId": f"A23.3-14-p{page:03d}-{per_page_counter[page]:02d}",
                "Page": page,
                "ImageNumber": num,
                "ImagePath": str(dest.relative_to(ROOT)),
                "WidthPx": row["width"],
                "HeightPx": row["height"],
                "ObjectId": row["object_id"],
                "CandidateKind": kind,
                "AssignedEquationLabel": assigned_label,
                "PageEquationLabels": "; ".join(labels),
                "PageContext": context_around_labels(page_text),
                "Status": "Needs review" if kind == "Equation candidate" else "Review",
                "SplitNeeded": "No",
                "PlainText": "",
                "LaTeX": "",
                "ExcelExpression": "",
                "Variables/Notes": "",
                "Reviewer": "",
            }
        )

    write_manifest(manifest_rows)
    write_workbook(manifest_rows)
    print(f"Extracted {len(manifest_rows)} embedded equation/image candidates")
    print(f"Manifest: {MANIFEST}")
    print(f"Workbook: {WORKBOOK}")
    print(f"Images:   {IMG_DIR}")


def write_manifest(rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    with MANIFEST.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def add_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("How to use")
    ws["A1"] = "A23.3-14 equation extraction review workbook"
    ws["A1"].font = Font(bold=True, size=14)
    guide = [
        ("Source of truth", "The embedded image is the source of truth. Transcribed text/LaTeX should be reviewed against the image."),
        ("Why this is cleaner", "This PDF stores equations as separate image objects, so the equation images are extracted intact instead of cropped from pages."),
        ("LaTeX", "Use LaTeX for exact math transcription, including subscripts, superscripts, roots, fractions, inequalities, and Greek letters."),
        ("PlainText", "Use only for search-friendly approximations."),
        ("ExcelExpression", "Optional. Use only if you want a calculable engineering expression."),
        ("SplitNeeded", "Set Yes when one image contains multiple equations or figure+equation content."),
        ("Status", "Needs review → Transcribed → Verified. Use Not equation for figures/logos/non-equation images."),
    ]
    for i, (k, v) in enumerate(guide, start=3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def write_workbook(rows: list[dict[str, object]]) -> None:
    headers = [
        "EquationId",
        "Page",
        "CandidateKind",
        "OriginalEquationImage",
        "ImagePath",
        "WidthPx",
        "HeightPx",
        "AssignedEquationLabel",
        "PageEquationLabels",
        "PageContext",
        "Status",
        "SplitNeeded",
        "PlainText",
        "LaTeX",
        "ExcelExpression",
        "Variables/Notes",
        "Reviewer",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Equation review"
    add_guide_sheet(wb)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_validation = DataValidation(
        type="list",
        formula1='"Needs review,Transcribed,Verified,Not equation,Needs split,Review"',
        allow_blank=False,
    )
    split_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(split_validation)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:Q1"

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row["EquationId"],
            row["Page"],
            row["CandidateKind"],
            None,
            row["ImagePath"],
            row["WidthPx"],
            row["HeightPx"],
            row["AssignedEquationLabel"],
            row["PageEquationLabels"],
            row["PageContext"],
            row["Status"],
            row["SplitNeeded"],
            row["PlainText"],
            row["LaTeX"],
            row["ExcelExpression"],
            row["Variables/Notes"],
            row["Reviewer"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        img_path = ROOT / str(row["ImagePath"])
        ws.cell(r_idx, 5).hyperlink = str(row["ImagePath"])
        ws.cell(r_idx, 5).style = "Hyperlink"
        status_validation.add(ws.cell(r_idx, 11))
        split_validation.add(ws.cell(r_idx, 12))

        if img_path.exists():
            with PilImage.open(img_path) as img:
                width, height = img.size
            max_width = 430
            max_height = 95
            scale = min(max_width / width, max_height / height, 1.0)
            ximg = XlsxImage(str(img_path))
            ximg.width = int(width * scale)
            ximg.height = int(height * scale)
            ws.add_image(ximg, f"D{r_idx}")
            ws.row_dimensions[r_idx].height = max(54, ximg.height * 0.82)

    widths = {
        "A": 22,
        "B": 8,
        "C": 18,
        "D": 60,
        "E": 54,
        "F": 9,
        "G": 9,
        "H": 24,
        "I": 30,
        "J": 80,
        "K": 16,
        "L": 12,
        "M": 38,
        "N": 52,
        "O": 42,
        "P": 44,
        "Q": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws["N1"].comment = Comment("Preferred exact transcription format.", "Codex")
    ws["O1"].comment = Comment("Optional calculable expression; do not use for non-computational formula text.", "Codex")
    wb.save(WORKBOOK)


if __name__ == "__main__":
    build_outputs()
