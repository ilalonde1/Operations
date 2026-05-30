"""
Extract docs/KOR Structural BD Tracking 2026.xlsx to a normalized JSONL.

One JSON object per spreadsheet row. The 6 regional sheets share an 11-column
shape (USA is the exception — it splits Location into City + State); we
normalize both into one object schema:

  {
    "region":       "Vancouver/LowerMainland" | "VancouverIsland" | "Alberta" |
                     "Okanagan-BcInterior" | "USA" | "EasternCanada",
    "rowNumber":   int (the spreadsheet "No." column, useful for traceability),
    "date":        "YYYY-MM-DD" or null,
    "initiator":   "Omar Alcazar" / "Conor Murtagh" / etc.,
    "contact":     person name (may be null),
    "company":     firm name (the Resolver target),
    "contactInfo": email/phone freeform,
    "location":    city as string (USA: "City, State"),
    "type":        "Meeting" | "Email" | "RFP" | etc.,
    "potentialProjects": freeform text (may be null),
    "proposalsSubmittedCad": decimal (USA submitted in USD; we convert below),
    "proposalsAcceptedCad":  decimal,
    "notes":       freeform,
    "currency":    "CAD" | "USD" (USA tab uses USD; CAD elsewhere)
  }

USA values are kept in source currency; the C# importer can convert via the
DASHBOARD's $US Conv. cell if a CAD figure is needed for rollups.

Output: tools/BdTrackingImport/inputs/bd-tracking.jsonl
"""
from openpyxl import load_workbook
from datetime import datetime, date as _date
import json
import warnings
import os

warnings.filterwarnings('ignore')

XLSX = r'C:/VIsual Studio Projects/Operations/docs/KOR Structural BD Tracking 2026.xlsx'
OUT_DIR = r'C:/VIsual Studio Projects/Operations/tools/BdTrackingImport/inputs'
OUT = os.path.join(OUT_DIR, 'bd-tracking.jsonl')

REGION_MAP = {
    'Vancouver; Lower Mainland': ('Vancouver/LowerMainland', 'CAD', False),
    'Vancouver Island':          ('VancouverIsland',         'CAD', False),
    'Alberta':                   ('Alberta',                 'CAD', False),
    'Okanagan; BC Interior':     ('Okanagan-BcInterior',     'CAD', False),
    'USA':                       ('USA',                     'USD', True),
    'Eastern Canada':            ('EasternCanada',           'CAD', False),
}

def safe_str(c):
    if c is None: return None
    s = str(c).strip()
    return s if s else None

def safe_decimal(c):
    if c is None: return None
    if isinstance(c, (int, float)):
        return float(c) if c != 0 else None
    s = str(c).strip().lower()
    if not s or s in ('lost', '-', '$0', '0', '$-'):
        return None
    # strip $ and commas
    s = s.replace('$','').replace(',','').strip()
    try:
        v = float(s)
        return v if v != 0 else None
    except ValueError:
        return None  # silently skip non-numeric

def safe_date(c):
    if c is None: return None
    if isinstance(c, datetime):
        return c.date().isoformat()
    if isinstance(c, _date):
        return c.isoformat()
    s = str(c).strip()
    if not s: return None
    # try YYYY-MM-DD parse
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date().isoformat()
    except ValueError:
        return None

def extract_sheet(ws, region_label, currency, has_city_state):
    rows = list(ws.iter_rows(values_only=True))
    # find header row (cell A = "No.")
    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == 'No.'), None)
    if header_idx is None:
        return []
    out = []
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None: continue
        # stop at TOTALS row
        if isinstance(r[0], str) and r[0].strip().upper().startswith('TOTAL'): break
        if not isinstance(r[0], (int, float)): continue  # skip stray text rows

        if has_city_state:
            # USA: No, Date, Initiator, Contact, Company, ContactInfo, City, State, Type, PotentialProjects, Submitted, Accepted, Notes
            row_no, dt, init, contact, company, cinfo, city, state, typ, potproj, sub, acc, notes = (r + (None,)*14)[:13]
            loc = ', '.join(p for p in [safe_str(city), safe_str(state)] if p)
            loc = loc if loc else None
        else:
            row_no, dt, init, contact, company, cinfo, loc, typ, potproj, sub, acc, notes = (r + (None,)*14)[:12]
            loc = safe_str(loc)

        rec = {
            'region':                region_label,
            'rowNumber':             int(row_no),
            'date':                  safe_date(dt),
            'initiator':             safe_str(init),
            'contact':               safe_str(contact),
            'company':               safe_str(company),
            'contactInfo':           safe_str(cinfo),
            'location':              loc,
            'type':                  safe_str(typ),
            'potentialProjects':     safe_str(potproj),
            'proposalsSubmittedCad': safe_decimal(sub),  # source currency on USA rows; importer reconverts
            'proposalsAcceptedCad':  safe_decimal(acc),
            'notes':                 safe_str(notes),
            'currency':              currency,
        }
        # skip rows that have nothing usable
        if not rec['initiator'] and not rec['company'] and not rec['contact']:
            continue
        out.append(rec)
    return out

def main():
    wb = load_workbook(XLSX, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)
    all_rows = []
    per_region = {}
    for sheet_name, (region_label, currency, has_city_state) in REGION_MAP.items():
        rows = extract_sheet(wb[sheet_name], region_label, currency, has_city_state)
        per_region[region_label] = len(rows)
        all_rows.extend(rows)

    with open(OUT, 'w', encoding='utf-8') as f:
        for r in all_rows:
            f.write(json.dumps(r, ensure_ascii=False) + '\n')

    print(f'Wrote {len(all_rows)} rows to {OUT}')
    for region, n in per_region.items():
        print(f'  {region:30s} {n:4d}')

if __name__ == '__main__':
    main()
